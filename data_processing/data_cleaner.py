import pandas as pd
from openpyxl import load_workbook

def extract_range_to_df(ws, cell_range):
    data = ws[cell_range]
    rows = [[cell.value for cell in row] for row in data]
    df = pd.DataFrame(rows[1:], columns=rows[0]) if len(rows) > 1 else pd.DataFrame(rows)
    return df

def clean_excel_file(input_file, output_file):
    wb = load_workbook(input_file, data_only=True)
    writer = pd.ExcelWriter(output_file, engine='openpyxl')

    if "Ladders" in wb.sheetnames:
        ws = wb["Ladders"]
        df = extract_range_to_df(ws, "B28:K35")
        if not df.empty:
            df.to_excel(writer, sheet_name="Ladders", index=False)

    if "Estimates" in wb.sheetnames:
        ws = wb["Estimates"]
        df = extract_range_to_df(ws, "A2:AD89")
        if not df.empty:
            df.to_excel(writer, sheet_name="Estimates", index=False)

    for sheet in ["Data Summary", "Water Data", "Sanitation Data", "Wastewater Data", "Hygiene Data", "Menstrual Health Data", "Population"]:
        if sheet in wb.sheetnames:
            df = pd.read_excel(input_file, sheet_name=sheet)
            if not df.empty:
                df.to_excel(writer, sheet_name=sheet[:31], index=False)

    writer.close()
    print(f"Data extracted and saved to: {output_file}")

if __name__ == "__main__":
    INPUT_PATH = "data/JMP_WASH_Data.xlsx"
    OUTPUT_PATH = "data/cleaned_Data.xlsx"
    clean_excel_file(INPUT_PATH, OUTPUT_PATH)